# -*- coding: utf-8 -*-
import os
import sys
sys.path.insert(0, '/usr/src/vekomet.ru/vekomet_redesign')
sys.path.append("/usr/src/vekomet.ru/vekomet_redesign")
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "vekomet_redesign.settings")
import django
django.setup()

from positions.models import Metall, PriceData
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Font
from openpyxl.drawing.image import Image

gr1=Metall.objects.filter(group=1)
gr2=Metall.objects.filter(group=2)
gr3=Metall.objects.filter(group=3)
gr4=Metall.objects.filter(group=4)
gr5=Metall.objects.filter(group=5)
gr6=Metall.objects.filter(group=6)
gr7=Metall.objects.filter(group=7)
gr8=Metall.objects.filter(group=8)
gr9=Metall.objects.filter(group=9)
gr10=Metall.objects.filter(group=10)
gr11=Metall.objects.filter(group=11)

wb = load_workbook('/usr/src/vekomet.ru/vekomet_redesign/static/core/images/PriceVekoStarter.xlsx')
ws = wb.active
try:
  os.remove('/usr/src/vekomet.ru/vekomet_redesign/static/core/images/PriceVeko.xlsx')
except:
  pass
fill1 = PatternFill(fill_type='solid',start_color='FFFFFFFF',end_color='FF000000')

border1 = Border(left=Side(border_style='thin',color='FF000000'),
                right=Side(border_style='thin',color='FF000000'),
                top=Side(border_style='thin',color='FF000000'),
                bottom=Side(border_style='thin',color='FF000000'))

border2 = Border(left=Side(border_style='thin',color='FF000000'),
                right=Side(border_style='thin',color='FF000000'),
                top=Side(border_style='thin',color='FF000000'),
                bottom=Side(border_style='thin',color='FF000000'))
font1 = Font(name='Calibri',size=14)
font2 = Font(name='Calibri',size=14, italic=True)

price=PriceData.objects.all()[0]

date=ws['B7']
date.value=price.pricedate
date.font=font1

count1=0
for s in gr1:
    count1+=1
counter1=11
for i in gr1:
  if i.visible:
    refB='B'+str(counter1)
    refC='C'+str(counter1)
    refD = 'D'+str(counter1)
    k=ws[refB]
    k.value=i.title
    k.border=border1
    k.font=font1
    l=ws[refC]
    l.value=i.price
    l.border=border1
    l.font=font1
    c=ws[refD]
    if i.price_bn:
        c.value=i.price_bn
    else:
        c.value = i.price
    c.border = border1
    c.font = font1
    counter1+=1
  else:
      pass
refadd1B='B'+str(counter1)
refadd1C='C'+str(counter1)
refadd1D='D'+str(counter1)
add1B=ws[refadd1B]
add1B.value=''
add1B.border=border2
add1B.fill=fill1
add1C=ws[refadd1C]
add1C.value=''
add1C.border=border2
add1C.fill=fill1
add1D=ws[refadd1D]
add1D.value=''
add1D.border=border2
add1D.fill=fill1

counter2=counter1+1
for i in gr2:
  if i.visible:
    refB='B'+str(counter2)
    refC='C'+str(counter2)
    refD ='D'+str(counter2)
    k=ws[refB]
    k.value=i.title
    k.border=border1
    k.font=font1
    l=ws[refC]
    l.value=i.price
    l.border=border1
    l.font=font1
    c = ws[refD]
    if i.price_bn:
        c.value = i.price_bn
    else:
        c.value = i.price
    c.border = border1
    c.font = font1
    counter2+=1
  else:
      pass
refadd2B='B'+str(counter2)
refadd2C='C'+str(counter2)
refadd12D='D'+str(counter2)
add2B=ws[refadd2B]
add2B.value=''
add2B.border=border2
add2B.fill=fill1
add2C=ws[refadd2C]
add2C.value=''
add2C.border=border2
add2C.fill=fill1
add12D=ws[refadd12D]
add12D.value=''
add12D.border=border2
add12D.fill=fill1

counter3=counter2+1
for i in gr3:
  if i.visible:
    refB='B'+str(counter3)
    refC='C'+str(counter3)
    refD = 'D' + str(counter3)
    k=ws[refB]
    k.value=i.title
    k.border=border1
    k.font=font1
    l=ws[refC]
    l.value=i.price
    l.border=border1
    l.font=font1
    c = ws[refD]
    if i.price_bn:
        c.value = i.price_bn
    else:
        c.value = i.price
    c.border = border1
    c.font = font1
    counter3+=1
  else:
      pass
refadd3B='B'+str(counter3)
refadd3C='C'+str(counter3)
refadd13D='D'+str(counter3)
add3B=ws[refadd3B]
add3B.value=''
add3B.border=border2
add3B.fill=fill1
add3C=ws[refadd3C]
add3C.value=''
add3C.border=border2
add3C.fill=fill1
add13D=ws[refadd13D]
add13D.value=''
add13D.border=border2
add13D.fill=fill1

counter4=counter3+1
for i in gr4:
  if i.visible:
    refB='B'+str(counter4)
    refC='C'+str(counter4)
    refD = 'D' + str(counter4)
    k=ws[refB]
    k.value=i.title
    k.border=border1
    k.font=font1
    l=ws[refC]
    l.value=i.price
    l.border=border1
    l.font=font1
    c = ws[refD]
    if i.price_bn:
        c.value = i.price_bn
    else:
        c.value = i.price
    c.border = border1
    c.font = font1
    counter4+=1
  else:
      pass
refadd4B='B'+str(counter4)
refadd4C='C'+str(counter4)
refadd14D='D'+str(counter4)
add4B=ws[refadd4B]
add4B.value=''
add4B.border=border2
add4B.fill=fill1
add4C=ws[refadd4C]
add4C.value=''
add4C.border=border2
add4C.fill=fill1
add14D=ws[refadd14D]
add14D.value=''
add14D.border=border2
add14D.fill=fill1

counter5=counter4+1
for i in gr5:
  if i.visible:
    refB='B'+str(counter5)
    refC='C'+str(counter5)
    refD = 'D' + str(counter5)
    k=ws[refB]
    k.value=i.title
    k.border=border1
    k.font=font1
    l=ws[refC]
    l.value=i.price
    l.border=border1
    l.font=font1
    c = ws[refD]
    if i.price_bn:
        c.value = i.price_bn
    else:
        c.value = i.price
    c.border = border1
    c.font = font1
    counter5+=1
  else:
      pass
refadd5B='B'+str(counter5)
refadd5C='C'+str(counter5)
refadd15D='D'+str(counter5)
add5B=ws[refadd5B]
add5B.value=''
add5B.border=border2
add5B.fill=fill1
add5C=ws[refadd5C]
add5C.value=''
add5C.border=border2
add5C.fill=fill1
add15D=ws[refadd15D]
add15D.value=''
add15D.border=border2
add15D.fill=fill1

counter6=counter5+1
for i in gr6:
  if i.visible:
    refB='B'+str(counter6)
    refC='C'+str(counter6)
    refD = 'D' + str(counter6)
    k=ws[refB]
    k.value=i.title
    k.border=border1
    k.font=font1
    l=ws[refC]
    l.value=i.price
    l.border=border1
    l.font=font1
    c = ws[refD]
    if i.price_bn:
        c.value = i.price_bn
    else:
        c.value = i.price
    c.border = border1
    c.font = font1
    counter6+=1
  else:
      pass
refadd6B='B'+str(counter6)
refadd6C='C'+str(counter6)
refadd16D='D'+str(counter6)
add6B=ws[refadd6B]
add6B.value=''
add6B.border=border2
add6B.fill=fill1
add6C=ws[refadd6C]
add6C.value=''
add6C.border=border2
add6C.fill=fill1
add16D=ws[refadd16D]
add16D.value=''
add16D.border=border2
add16D.fill=fill1

counter7=counter6+1
for i in gr7:
  if i.visible:
    refB='B'+str(counter7)
    refC='C'+str(counter7)
    refD = 'D' + str(counter7)
    k=ws[refB]
    k.value=i.title
    k.border=border1
    k.font=font1
    l=ws[refC]
    l.value=i.price
    l.border=border1
    l.font=font1
    c = ws[refD]
    if i.price_bn:
        c.value = i.price_bn
    else:
        c.value = i.price
    c.border = border1
    c.font = font1
    counter7+=1
  else:
      pass
refadd7B='B'+str(counter7)
refadd7C='C'+str(counter7)
refadd17D='D'+str(counter7)
add7B=ws[refadd7B]
add7B.value=''
add7B.border=border2
add7B.fill=fill1
add7C=ws[refadd7C]
add7C.value=''
add7C.border=border2
add7C.fill=fill1
add17D=ws[refadd17D]
add17D.value=''
add17D.border=border2
add17D.fill=fill1

counter8=counter7+1
for i in gr8:
  if i.visible:
    refB='B'+str(counter8)
    refC='C'+str(counter8)
    refD = 'D' + str(counter8)
    k=ws[refB]
    k.value=i.title
    k.border=border1
    k.font=font1
    l=ws[refC]
    l.value=i.price
    l.border=border1
    l.font=font1
    c = ws[refD]
    if i.price_bn:
        c.value = i.price_bn
    else:
        c.value = i.price
    c.border = border1
    c.font = font1
    counter8+=1
  else:
      pass
refadd8B='B'+str(counter8)
refadd8C='C'+str(counter8)
refadd18D='D'+str(counter8)
add8B=ws[refadd8B]
add8B.value=''
add8B.border=border2
add8B.fill=fill1
add8C=ws[refadd8C]
add8C.value=''
add8C.border=border2
add8C.fill=fill1
add18D=ws[refadd18D]
add18D.value=''
add18D.border=border2
add18D.fill=fill1

counter9=counter8+1
for i in gr9:
  if i.visible:
    refB='B'+str(counter9)
    refC='C'+str(counter9)
    refD = 'D' + str(counter9)
    k=ws[refB]
    k.value=i.title
    k.border=border1
    k.font=font1
    l=ws[refC]
    l.value=i.price
    l.border=border1
    l.font=font1
    c = ws[refD]
    if i.price_bn:
        c.value = i.price_bn
    else:
        c.value = i.price
    c.border = border1
    c.font = font1
    counter9+=1
  else:
      pass
refadd9B='B'+str(counter9)
refadd9C='C'+str(counter9)
refadd19D='D'+str(counter9)
add9B=ws[refadd9B]
add9B.value=''
add9B.border=border2
add9B.fill=fill1
add9C=ws[refadd9C]
add9C.value=''
add9C.border=border2
add9C.fill=fill1
add19D=ws[refadd19D]
add19D.value=''
add19D.border=border2
add19D.fill=fill1

counter10=counter9+1
for i in gr10:
  if i.visible:
    refB='B'+str(counter10)
    refC='C'+str(counter10)
    refD = 'D' + str(counter10)
    k=ws[refB]
    k.value=i.title
    k.border=border1
    k.font=font1
    l=ws[refC]
    l.value=i.price
    l.border=border1
    l.font=font1
    c = ws[refD]
    if i.price_bn:
        c.value = i.price_bn
    else:
        c.value = i.price
    c.border = border1
    c.font = font1
    counter10+=1
  else:
      pass
refadd10B='B'+str(counter10)
refadd10C='C'+str(counter10)
refadd110D='D'+str(counter10)
add10B=ws[refadd10B]
add10B.value=''
add10B.border=border2
add10B.fill=fill1
add10C=ws[refadd10C]
add10C.value=''
add10C.border=border2
add10C.fill=fill1
add110D=ws[refadd1D]
add110D.value=''
add110D.border=border2
add110D.fill=fill1

counter11=counter10+1
for i in gr11:
  if i.visible:
    refB='B'+str(counter11)
    refC='C'+str(counter11)
    refD = 'D' + str(counter11)
    k=ws[refB]
    k.value=i.title
    k.border=border1
    k.font=font1
    l=ws[refC]
    l.value=i.price
    l.border=border1
    l.font=font1
    c = ws[refD]
    if i.price_bn:
        c.value = i.price_bn
    else:
        c.value = i.price
    c.border = border1
    c.font = font1
    counter11+=1
  else:
      pass
refadd11B='B'+str(counter11)
refadd11C='C'+str(counter11)
refadd111D='D'+str(counter11)
add11B=ws[refadd11B]
add11B.value=''
add11B.border=border2
add11B.fill=fill1
add11C=ws[refadd11C]
add11C.value=''
add11C.border=border2
add11C.fill=fill1
add111D=ws[refadd111D]
add111D.value=''
add111D.border=border2
add111D.fill=fill1


img = Image('/usr/src/vekomet.ru/vekomet_redesign/static/core/images/logo.png')
ws.add_image(img, 'C1')

wb.save('/usr/src/vekomet.ru/vekomet_redesign/static/core/images/PriceVeko.xlsx')
print ('excellent')